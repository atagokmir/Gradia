import io
from datetime import datetime

import openpyxl
from django.http import HttpResponse

from .models import Group, Rating, Student, Survey


def import_students_from_excel(file):
    """
    Import students from an Excel file.
    Expected columns: ad_soyad, kullanici_adi, ogrenci_no, grup
    Returns dict with added_count, error_count, error_details.
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    added_count = 0
    error_count = 0
    error_details = []

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

    col_map = {}
    for idx, h in enumerate(headers):
        if h in ('ad_soyad', 'kullanici_adi', 'ogrenci_no', 'grup'):
            col_map[h] = idx

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            ad_soyad = str(row[col_map['ad_soyad']] or '').strip()
            kullanici_adi = str(row[col_map['kullanici_adi']] or '').strip()
            ogrenci_no = str(row[col_map['ogrenci_no']] or '').strip()
            grup = str(row[col_map['grup']] or '').strip()

            if not kullanici_adi:
                error_count += 1
                error_details.append(f"Satır {row_num}: Kullanıcı adı boş")
                continue

            # Get or create group
            group = None
            if grup:
                group, _ = Group.objects.get_or_create(name=grup)

            # Skip if student with username exists
            if Student.objects.filter(username=kullanici_adi).exists():
                error_count += 1
                error_details.append(f"Satır {row_num}: '{kullanici_adi}' zaten mevcut")
                continue

            # Parse first/last name
            parts = ad_soyad.split(' ', 1)
            first_name = parts[0] if parts else ''
            last_name = parts[1] if len(parts) > 1 else ''

            student = Student(
                username=kullanici_adi,
                first_name=first_name,
                last_name=last_name,
                ogrenci_no=ogrenci_no,
                group=group,
            )
            student.set_password(ogrenci_no)
            student.save()
            added_count += 1

        except (KeyError, TypeError, IndexError) as e:
            error_count += 1
            error_details.append(f"Satır {row_num}: {str(e)}")

    return {
        'added_count': added_count,
        'error_count': error_count,
        'error_details': error_details,
    }


def export_summary_excel(survey):
    """Export summary: one row per student with average score."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Özet'

    headers = ['Ad Soyad', 'Kullanıcı Adı', 'Öğrenci No', 'Grup', 'Ortalama Puan', 'Değerlendiren Sayısı']
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    ratings = Rating.objects.filter(survey=survey).select_related('ratee', 'ratee__group')
    ratee_ids = ratings.values_list('ratee_id', flat=True).distinct()

    students = Student.objects.filter(id__in=ratee_ids).select_related('group').order_by('group__name', 'last_name', 'first_name')

    for student in students:
        student_ratings = ratings.filter(ratee=student)
        scores = [r.score for r in student_ratings]
        avg = round(sum(scores) / len(scores), 2) if scores else 0
        ws.append([
            student.get_full_name(),
            student.username,
            student.ogrenci_no,
            student.group.name if student.group else '',
            avg,
            len(scores),
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="ozet_{survey.id}.xlsx"'
    return response


def export_detail_excel(survey):
    """Export detail: one row per rating."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detay'

    headers = ['Değerlendiren', 'Değerlendirilen', 'Puan', 'Yorum', 'Tarih']
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    ratings = Rating.objects.filter(survey=survey).select_related('rater', 'ratee').order_by('rater__last_name', 'rater__first_name')

    for rating in ratings:
        ws.append([
            rating.rater.get_full_name(),
            rating.ratee.get_full_name(),
            rating.score,
            rating.comment,
            rating.created_at.strftime('%Y-%m-%d %H:%M') if rating.created_at else '',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="detay_{survey.id}.xlsx"'
    return response
